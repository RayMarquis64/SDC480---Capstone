"""
Excel export functionality for Project Pricer
Requires: pip install openpyxl
"""
from datetime import datetime

def export_project_to_excel(cursor, project_id, filename):
    """
    Export a project to Excel format
    
    Args:
        cursor: SQLite cursor
        project_id: ID of project to export
        filename: Path to save Excel file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Summary"
    
    # Get project details
    cursor.execute('''
        SELECT p.name, p.description, p.created_date, pr.name, pr.hourly_rate
        FROM projects p
        JOIN profiles pr ON p.profile_id = pr.id
        WHERE p.id = ?
    ''', (project_id,))
    project = cursor.fetchone()
    
    if not project:
        raise ValueError("Project not found")
    
    project_name, description, created_date, profile_name, hourly_rate = project
    
    # Styles
    title_font = Font(size=16, bold=True)
    header_font = Font(size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subtotal_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    total_font = Font(size=12, bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    row = 1
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = f"PROJECT COST ESTIMATE: {project_name}"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center')
    
    # Project info
    row += 2
    ws[f'A{row}'] = "Profile:"
    ws[f'B{row}'] = profile_name
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Description:"
    ws[f'B{row}'] = description
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Created Date:"
    ws[f'B{row}'] = created_date[:10]
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Hourly Rate:"
    ws[f'B{row}'] = f"${hourly_rate:.2f}"
    ws[f'A{row}'].font = Font(bold=True)
    
    # Materials section
    row += 3
    ws[f'A{row}'] = "MATERIALS"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    
    row += 1
    headers = ['Item', 'Quantity', 'Unit Cost', 'Total Cost']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Get materials
    cursor.execute('''
        SELECT name, quantity, unit_cost
        FROM materials
        WHERE project_id = ?
    ''', (project_id,))
    materials = cursor.fetchall()
    
    materials_total = 0
    for mat in materials:
        row += 1
        total = (mat[1] or 0) * (mat[2] or 0)
        materials_total += total
        
        ws[f'A{row}'] = mat[0]
        ws[f'B{row}'] = mat[1]
        ws[f'C{row}'] = f"${mat[2]:.2f}"
        ws[f'D{row}'] = f"${total:.2f}"
        
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
    
    # Materials subtotal
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = "MATERIALS SUBTOTAL"
    cell.font = Font(bold=True)
    cell.fill = subtotal_fill
    cell.alignment = Alignment(horizontal='right')
    
    cell = ws[f'D{row}']
    cell.value = f"${materials_total:.2f}"
    cell.font = Font(bold=True)
    cell.fill = subtotal_fill
    
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = border
    
    # Labor section
    row += 3
    ws[f'A{row}'] = "LABOR"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    
    row += 1
    headers = ['Description', 'Hours', 'Rate', 'Total Cost']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Get labor
    cursor.execute('''
        SELECT description, hours
        FROM labor
        WHERE project_id = ?
    ''', (project_id,))
    labor = cursor.fetchall()
    
    labor_total = 0
    for lab in labor:
        row += 1
        total = (lab[1] or 0) * hourly_rate
        labor_total += total
        
        ws[f'A{row}'] = lab[0]
        ws[f'B{row}'] = lab[1]
        ws[f'C{row}'] = f"${hourly_rate:.2f}/hr"
        ws[f'D{row}'] = f"${total:.2f}"
        
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
    
    # Labor subtotal
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = "LABOR SUBTOTAL"
    cell.font = Font(bold=True)
    cell.fill = subtotal_fill
    cell.alignment = Alignment(horizontal='right')
    
    cell = ws[f'D{row}']
    cell.value = f"${labor_total:.2f}"
    cell.font = Font(bold=True)
    cell.fill = subtotal_fill
    
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = border
    
    # Tool usage section
    row += 3
    ws[f'A{row}'] = "TOOL USAGE"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    
    row += 1
    headers = ['Tool/Machine', 'Hours', 'Rate', 'Total Cost']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Get tool usage
    cursor.execute('''
        SELECT t.name, tu.hours, t.cost_per_hour
        FROM tool_usage tu
        JOIN tools t ON tu.tool_id = t.id
        WHERE tu.project_id = ?
    ''', (project_id,))
    tools = cursor.fetchall()
    
    tools_total = 0
    for tool in tools:
        row += 1
        total = (tool[1] or 0) * (tool[2] or 0)
        tools_total += total
        
        ws[f'A{row}'] = tool[0]
        ws[f'B{row}'] = tool[1]
        ws[f'C{row}'] = f"${tool[2]:.2f}/hr"
        ws[f'D{row}'] = f"${total:.2f}"
        
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
    
    # Tools subtotal
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = "TOOL USAGE SUBTOTAL"
    cell.font = Font(bold=True)
    cell.fill = subtotal_fill
    cell.alignment = Alignment(horizontal='right')
    
    cell = ws[f'D{row}']
    cell.value = f"${tools_total:.2f}"
    cell.font = Font(bold=True)
    cell.fill = subtotal_fill
    
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = border
    
    # Grand total
    row += 2
    grand_total = materials_total + labor_total + tools_total
    
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = "GRAND TOTAL"
    cell.font = total_font
    cell.fill = total_fill
    cell.alignment = Alignment(horizontal='right')
    
    cell = ws[f'D{row}']
    cell.value = f"${grand_total:.2f}"
    cell.font = total_font
    cell.fill = total_fill
    
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Save workbook
    wb.save(filename)
    return filename